import openpyxl
import sqlite3

EXCEL_FILE = r"C:\Users\Lalitha\OneDrive\Desktop\Middle School\Middle_School_Olympiad_Platform_BIDMAS.xlsx"
DB_FILE = r"C:\Users\Lalitha\OneDrive\Desktop\Middle School\olympiad_questions.db"

# Load Excel workbook
wb = openpyxl.load_workbook(EXCEL_FILE)
sheet = wb.active

# Get headers
headers = {cell.value: col for col, cell in enumerate(sheet[1], 1)}
question_col = headers.get('Question')
exam_col = headers.get('Exam')

# Connect to SQLite
conn = sqlite3.connect(DB_FILE)
cursor = conn.cursor()

updated = 0

# Update rows 74-77
for row_num in range(74, 78):
    question = sheet.cell(row=row_num, column=question_col).value
    exam = sheet.cell(row=row_num, column=exam_col).value

    if question and exam:
        question = str(question).strip()
        exam = str(exam).strip()

        cursor.execute(
            """
            UPDATE questions
            SET exam = ?
            WHERE question = ?
            """,
            (exam, question)
        )

        updated += cursor.rowcount
        print(f"Row {row_num}: Updated question with exam = {exam}")

conn.commit()
conn.close()

print(f"\nTotal updated: {updated} rows.")
